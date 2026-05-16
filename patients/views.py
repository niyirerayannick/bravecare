from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q
from django.core.paginator import Paginator
from .models import Patient, Screening, MaternalChildHealth, FollowUp
from .forms import PatientForm, FollowUpForm, ScreeningForm
from locations.models import Province, District, Sector, Cell, Village


def _is_ajax(request):
    return request.headers.get('X-Requested-With') == 'XMLHttpRequest'


def _form_errors(form):
    return {k: [str(e) for e in v] for k, v in form.errors.items()}


# ─────────────────────── PATIENTS ───────────────────────

@login_required
def patient_list(request):
    query           = request.GET.get('q', '')
    risk_filter     = request.GET.get('risk', '')
    province_filter = request.GET.get('province', '')
    district_filter = request.GET.get('district', '')
    sector_filter   = request.GET.get('sector', '')
    cell_filter     = request.GET.get('cell', '')
    village_filter  = request.GET.get('village', '')
    patients = Patient.objects.all()

    if query:
        patients = patients.filter(
            Q(first_name__icontains=query) | Q(last_name__icontains=query) |
            Q(phone__icontains=query) | Q(national_id__icontains=query)
        )
    if risk_filter:     patients = patients.filter(risk_level=risk_filter)
    if province_filter: patients = patients.filter(province_id=province_filter)
    if district_filter: patients = patients.filter(district_id=district_filter)
    if sector_filter:   patients = patients.filter(sector_id=sector_filter)
    if cell_filter:     patients = patients.filter(cell_id=cell_filter)
    if village_filter:  patients = patients.filter(village_id=village_filter)

    # Each level scoped to its parent selection for the filter dropdowns
    districts = District.objects.filter(province_id=province_filter).order_by('name') if province_filter else District.objects.order_by('name')
    sectors   = Sector.objects.filter(district_id=district_filter).order_by('name')   if district_filter else Sector.objects.none()
    cells     = Cell.objects.filter(sector_id=sector_filter).order_by('name')         if sector_filter   else Cell.objects.none()
    villages  = Village.objects.filter(cell_id=cell_filter).order_by('name')          if cell_filter     else Village.objects.none()

    context = {
        'patients': patients,
        'query': query,
        'risk_filter':     risk_filter,
        'province_filter': province_filter,
        'district_filter': district_filter,
        'sector_filter':   sector_filter,
        'cell_filter':     cell_filter,
        'village_filter':  village_filter,
        'provinces': Province.objects.order_by('name'),
        'districts': districts,
        'sectors':   sectors,
        'cells':     cells,
        'villages':  villages,
        'total_count':       Patient.objects.count(),
        'high_risk_count':   Patient.objects.filter(risk_level='high').count(),
        'medium_risk_count': Patient.objects.filter(risk_level='medium').count(),
        'low_risk_count':    Patient.objects.filter(risk_level='low').count(),
    }
    return render(request, 'patients/list.html', context)


@login_required
def patient_detail(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    context = {
        'patient': patient,
        'screenings': patient.screenings.all()[:10],
        'followups': patient.followups.all()[:5],
        'mch_records': patient.mch_records.all()[:5],
        'all_patients': Patient.objects.all().order_by('first_name', 'last_name'),
        'screening_type_choices': Screening.SCREENING_TYPE_CHOICES,
    }
    return render(request, 'patients/detail.html', context)


def _apply_location(obj, post):
    """Copy location FK IDs and address_text from POST data onto a model instance."""
    for field in ('country_id', 'province_id', 'district_id', 'sector_id', 'cell_id', 'village_id'):
        val = post.get(field, '').strip()
        setattr(obj, field, int(val) if val.isdigit() else None)
    obj.address_text = post.get('address_text', '').strip()


def _location_json(obj):
    return {
        'country_id':  obj.country_id,
        'province_id': obj.province_id,
        'district_id': obj.district_id,
        'sector_id':   obj.sector_id,
        'cell_id':     obj.cell_id,
        'village_id':  obj.village_id,
        'address_text': obj.address_text,
        'display_location': obj.display_location(),
    }


@login_required
def patient_add(request):
    if request.method == 'POST':
        form = PatientForm(request.POST)
        if form.is_valid():
            patient = form.save(commit=False)
            _apply_location(patient, request.POST)
            patient.save()
            if _is_ajax(request):
                return JsonResponse({'success': True, 'message': f'{patient.get_full_name()} registered.'})
            messages.success(request, f'Patient {patient.get_full_name()} added.')
            return redirect('patients:list')
        if _is_ajax(request):
            return JsonResponse({'success': False, 'errors': _form_errors(form)}, status=400)
    form = PatientForm()
    return render(request, 'patients/form.html', {'form': form, 'title': 'Add New Patient'})


@login_required
def patient_edit(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    if request.method == 'POST':
        form = PatientForm(request.POST, instance=patient)
        if form.is_valid():
            patient = form.save(commit=False)
            _apply_location(patient, request.POST)
            patient.save()
            if _is_ajax(request):
                return JsonResponse({'success': True, 'message': f'{patient.get_full_name()} updated.'})
            messages.success(request, 'Patient updated.')
            return redirect('patients:detail', pk=patient.pk)
        if _is_ajax(request):
            return JsonResponse({'success': False, 'errors': _form_errors(form)}, status=400)
    form = PatientForm(instance=patient)
    return render(request, 'patients/form.html', {'form': form, 'title': 'Edit Patient', 'patient': patient})


@login_required
def patient_json(request, pk):
    p = get_object_or_404(Patient, pk=pk)
    data = {
        'first_name': p.first_name,
        'last_name': p.last_name,
        'phone': p.phone,
        'national_id': p.national_id,
        'date_of_birth': str(p.date_of_birth),
        'gender': p.gender,
        'risk_level': p.risk_level,
        'notes': p.notes,
    }
    data.update(_location_json(p))
    return JsonResponse(data)


# ─────────────────────── SCREENINGS ───────────────────────

def _screening_queryset(request):
    """Return filtered Screening queryset based on GET params (shared by list and export)."""
    query = request.GET.get('q', '')
    type_filter = request.GET.get('type', '')
    result_filter = request.GET.get('result', '')
    status_filter = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    qs = Screening.objects.select_related('patient', 'performed_by').all()

    if query:
        qs = qs.filter(
            Q(patient__first_name__icontains=query) |
            Q(patient__last_name__icontains=query) |
            Q(patient__phone__icontains=query) |
            Q(patient__national_id__icontains=query) |
            Q(screening_type__icontains=query)
        )
    if type_filter:
        qs = qs.filter(screening_type=type_filter)
    if result_filter:
        qs = qs.filter(result__icontains=result_filter)
    if status_filter:
        qs = qs.filter(status=status_filter)
    if date_from:
        qs = qs.filter(screening_date__gte=date_from)
    if date_to:
        qs = qs.filter(screening_date__lte=date_to)
    return qs


@login_required
def screening_list(request):
    query = request.GET.get('q', '')
    type_filter = request.GET.get('type', '')
    result_filter = request.GET.get('result', '')
    status_filter = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    screenings_qs = _screening_queryset(request)
    paginator = Paginator(screenings_qs, 25)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    total = Screening.objects.count()
    pending = Screening.objects.filter(status='pending').count()
    referred = Screening.objects.filter(status='referred').count()
    positive = Screening.objects.filter(
        Q(result__icontains='positive') | Q(result__icontains='diabetic') |
        Q(result__icontains='critical') | Q(result__icontains='severe') |
        Q(result__icontains='confirmed') | Q(result__icontains='abnormal')
    ).count()

    context = {
        'page_obj': page_obj,
        'query': query,
        'type_filter': type_filter,
        'result_filter': result_filter,
        'status_filter': status_filter,
        'date_from': date_from,
        'date_to': date_to,
        'total_count': total,
        'pending_count': pending,
        'positive_count': positive,
        'referred_count': referred,
        'screening_type_choices': Screening.SCREENING_TYPE_CHOICES,
        'all_patients': Patient.objects.all().order_by('first_name', 'last_name'),
        # keep legacy key for any remaining references
        'screenings': page_obj,
        'active_count': pending,
    }
    return render(request, 'patients/screening_list.html', context)


@login_required
def patient_autocomplete(request):
    q = request.GET.get('q', '').strip()
    patients = Patient.objects.all().order_by('first_name', 'last_name')
    if q:
        patients = patients.filter(
            Q(first_name__icontains=q) | Q(last_name__icontains=q) |
            Q(phone__icontains=q) | Q(national_id__icontains=q)
        )
    results = [
        {'id': p.pk, 'text': f"{p.get_full_name()} — {p.phone} (ID: {p.national_id})"}
        for p in patients[:50]
    ]
    return JsonResponse({'results': results})


@login_required
def screening_add(request):
    if request.method == 'POST':
        form = ScreeningForm(request.POST)
        if form.is_valid():
            s = form.save(commit=False)
            s.performed_by = request.user
            s.save()
            if _is_ajax(request):
                return JsonResponse({'success': True, 'message': 'Screening record added.'})
            messages.success(request, 'Screening added.')
            return redirect('patients:screening_list')
        if _is_ajax(request):
            return JsonResponse({'success': False, 'errors': _form_errors(form)}, status=400)
    return redirect('patients:screening_list')


@login_required
def screening_edit(request, pk):
    s = get_object_or_404(Screening, pk=pk)
    if request.method == 'POST':
        form = ScreeningForm(request.POST, instance=s)
        if form.is_valid():
            form.save()
            if _is_ajax(request):
                return JsonResponse({'success': True, 'message': 'Screening updated.'})
            messages.success(request, 'Screening updated.')
            return redirect('patients:screening_list')
        if _is_ajax(request):
            return JsonResponse({'success': False, 'errors': _form_errors(form)}, status=400)
    return redirect('patients:screening_list')


@login_required
def screening_json(request, pk):
    s = get_object_or_404(Screening, pk=pk)
    return JsonResponse({
        'patient': s.patient_id,
        'patient_text': f"{s.patient.get_full_name()} — {s.patient.phone} (ID: {s.patient.national_id})",
        'screening_type': s.screening_type,
        'result': s.result,
        'risk_level': s.risk_level,
        'screening_date': str(s.screening_date),
        'status': s.status,
        'comments': s.comments,
        'recommendations': s.recommendations,
        'followup_notes': s.followup_notes,
        'next_followup_date': str(s.next_followup_date) if s.next_followup_date else '',
        'notes': s.notes,
    })


@login_required
def screening_export_csv(request):
    import csv
    qs = _screening_queryset(request)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="screenings.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'Patient', 'National ID', 'Phone', 'Screening Type', 'Result',
        'Risk Level', 'Status', 'Screening Date', 'Next Follow-up',
        'Comments', 'Recommendations', 'Follow-up Notes', 'Healthcare Worker',
    ])
    for s in qs:
        writer.writerow([
            s.patient.get_full_name(),
            s.patient.national_id,
            s.patient.phone,
            s.get_screening_type_display(),
            s.result,
            s.get_risk_level_display(),
            s.get_status_display(),
            str(s.screening_date),
            str(s.next_followup_date) if s.next_followup_date else '',
            s.comments,
            s.recommendations,
            s.followup_notes,
            s.performed_by.get_full_name() if s.performed_by else '',
        ])
    return response


@login_required
def screening_export_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        qs = _screening_queryset(request)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Screenings'

        headers = [
            'Patient', 'National ID', 'Phone', 'Screening Type', 'Result',
            'Risk Level', 'Status', 'Screening Date', 'Next Follow-up',
            'Comments', 'Recommendations', 'Follow-up Notes', 'Healthcare Worker',
        ]
        header_fill = PatternFill('solid', fgColor='10B981')
        header_font = Font(bold=True, color='FFFFFF')
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for s in qs:
            ws.append([
                s.patient.get_full_name(),
                s.patient.national_id,
                s.patient.phone,
                s.get_screening_type_display(),
                s.result,
                s.get_risk_level_display(),
                s.get_status_display(),
                str(s.screening_date),
                str(s.next_followup_date) if s.next_followup_date else '',
                s.comments,
                s.recommendations,
                s.followup_notes,
                s.performed_by.get_full_name() if s.performed_by else '',
            ])

        for column in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in column), default=10)
            ws.column_dimensions[column[0].column_letter].width = min(max_len + 4, 40)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="screenings.xlsx"'
        wb.save(response)
        return response
    except ImportError:
        return screening_export_csv(request)


@login_required
def screening_export_pdf(request):
    """Print-friendly HTML view that the browser can print/save as PDF."""
    qs = _screening_queryset(request)
    return render(request, 'patients/screening_print.html', {
        'screenings': qs,
        'query': request.GET.get('q', ''),
    })


# ─────────────────────── FOLLOW-UPS ───────────────────────

@login_required
def followup_list(request):
    status_filter = request.GET.get('status', '')
    followups = FollowUp.objects.select_related('patient').all()

    if status_filter:
        followups = followups.filter(status=status_filter)

    context = {
        'followups': followups,
        'status_filter': status_filter,
        'all_patients': Patient.objects.all().order_by('first_name', 'last_name'),
    }
    return render(request, 'patients/followup_list.html', context)


@login_required
def followup_add(request):
    if request.method == 'POST':
        form = FollowUpForm(request.POST)
        if form.is_valid():
            f = form.save(commit=False)
            f.assigned_to = request.user
            f.save()
            if _is_ajax(request):
                return JsonResponse({'success': True, 'message': 'Follow-up scheduled.'})
            messages.success(request, 'Follow-up added.')
            return redirect('patients:followup_list')
        if _is_ajax(request):
            return JsonResponse({'success': False, 'errors': _form_errors(form)}, status=400)
    form = FollowUpForm()
    return render(request, 'patients/form.html', {'form': form, 'title': 'Add Follow-up'})


@login_required
def followup_edit(request, pk):
    fu = get_object_or_404(FollowUp, pk=pk)
    if request.method == 'POST':
        form = FollowUpForm(request.POST, instance=fu)
        if form.is_valid():
            form.save()
            if _is_ajax(request):
                return JsonResponse({'success': True, 'message': 'Follow-up updated.'})
            messages.success(request, 'Follow-up updated.')
            return redirect('patients:followup_list')
        if _is_ajax(request):
            return JsonResponse({'success': False, 'errors': _form_errors(form)}, status=400)
    form = FollowUpForm(instance=fu)
    return render(request, 'patients/form.html', {'form': form, 'title': 'Edit Follow-up'})


@login_required
def followup_json(request, pk):
    f = get_object_or_404(FollowUp, pk=pk)
    return JsonResponse({
        'patient': f.patient_id,
        'service': f.service,
        'due_date': str(f.due_date),
        'status': f.status,
        'notes': f.notes,
    })


# ─────────────────────── MCH ───────────────────────

@login_required
def mch_add(request):
    from django.forms import ModelForm

    class MCHForm(ModelForm):
        class Meta:
            from .models import MaternalChildHealth
            model = MaternalChildHealth
            fields = ['patient', 'service_type', 'visit_date', 'next_followup_date', 'notes']

    if request.method == 'POST':
        form = MCHForm(request.POST)
        if form.is_valid():
            m = form.save(commit=False)
            m.attended_by = request.user
            m.save()
            if _is_ajax(request):
                return JsonResponse({'success': True, 'message': 'MCH record added.'})
            messages.success(request, 'MCH record added.')
            return redirect('patients:list')
        if _is_ajax(request):
            return JsonResponse({'success': False, 'errors': _form_errors(form)}, status=400)
    return JsonResponse({'error': 'POST only'}, status=405)


@login_required
def mch_edit(request, pk):
    from django.forms import ModelForm

    class MCHForm(ModelForm):
        class Meta:
            from .models import MaternalChildHealth
            model = MaternalChildHealth
            fields = ['patient', 'service_type', 'visit_date', 'next_followup_date', 'notes']

    mch = get_object_or_404(MaternalChildHealth, pk=pk)
    if request.method == 'POST':
        form = MCHForm(request.POST, instance=mch)
        if form.is_valid():
            form.save()
            if _is_ajax(request):
                return JsonResponse({'success': True, 'message': 'MCH record updated.'})
            messages.success(request, 'MCH record updated.')
        if _is_ajax(request):
            return JsonResponse({'success': False, 'errors': _form_errors(form)}, status=400)
    return JsonResponse({'error': 'POST only'}, status=405)


@login_required
def mch_json(request, pk):
    m = get_object_or_404(MaternalChildHealth, pk=pk)
    return JsonResponse({
        'patient': m.patient_id,
        'service_type': m.service_type,
        'visit_date': str(m.visit_date),
        'next_followup_date': str(m.next_followup_date) if m.next_followup_date else '',
        'notes': m.notes,
    })
